import re
import chardet
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

# File paths
log_file = "/Users/aliesmaeilicharkhab/Documents/PersonalProjects/Python/RSL/RSL.log"
reservation_file = "/Users/aliesmaeilicharkhab/Documents/PersonalProjects/Python/RSL/port_reservation.xlsx"
with open(log_file, "rb") as f:
    raw = f.read()
    enc = chardet.detect(raw)["encoding"] or "utf-8"

# --- Regex for optics data (allow N/A, + in type) ---
pattern = re.compile(
    r"Optics\s+([\d/]+)\s+"              # Port (e.g., 0/0/0/6)
    r"\w+\s+\w+\s+\w+\s+\w+\s+\d+\s+[\d.]+mA\s+"  # Skip fields up to Laser Bias
    r"(-?\d+\.\d+|N/?A)\s+"              # TX Power
    r"(-?\d+\.\d+|N/?A)\s+"              # RX Power
    r"([\d.]+)\s+"                       # Wavelength (ignored)
    r"([A-Za-z0-9+\-_/]+)",              # Optics Type (allows + - _ /)
    flags=re.IGNORECASE
)

# --- Extract optics info from log into dict keyed by "0/0/0/6" ---
optics = {}
for line in raw.decode(enc, errors="ignore").splitlines():
    m = pattern.search(line)
    if not m:
        continue
    port = m.group(1)
    tx = m.group(2)
    rx = m.group(3)
    typ = m.group(5)
    optics[port] = {"tx": tx, "rx": rx, "type": typ}

# --- Helpers ---
def is_na(val: str) -> bool:
    s = str(val).strip().upper()
    return s in {"N/A", "NA", "#N/A"}

def set_cell_numeric_or_na(ws, addr: str, value):
    """Write float when possible; write 'N/A' exactly when source is N/A; leave blank if unparsable."""
    s = "" if value is None else str(value).strip()
    if is_na(s):
        ws[addr] = "N/A"
        return
    try:
        num = float(s)
        ws[addr] = num
        ws[addr].number_format = "0.00"
    except ValueError:
        ws[addr] = None  # leave blank if unexpected text

# --- Load workbook & first sheet ---
wb = load_workbook(reservation_file)
ws = wb.active

# --- Fill columns U (TX), V (RX), W (Type) by matching Column E's interface ---
# Column E contains names like GigabitEthernet0/0/0/6 or TenGigE0/0/0/6 → extract 0/0/0/6 and match.
maxr = ws.max_row
for r in range(2, maxr + 1):
    name = ws[f"E{r}"].value
    if not name:
        continue
    m = re.search(r"(\d+/\d+/\d+/\d+)", str(name))
    if not m:
        continue
    key = m.group(1)  # e.g., 0/0/0/6
    if key not in optics:
        continue
    # Write TX (U), RX (V), Type (W)
    set_cell_numeric_or_na(ws, f"U{r}", optics[key]["tx"])
    set_cell_numeric_or_na(ws, f"V{r}", optics[key]["rx"])
    ws[f"W{r}"] = optics[key]["type"]

# --- Conditional formatting (numeric-only) ---
red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

u_range = f"U2:U{maxr}"
v_range = f"V2:V{maxr}"

# RED: value < -18 (numbers only)
ws.conditional_formatting.add(
    u_range, FormulaRule(formula=["AND(ISNUMBER(U2),U2<-18)"], fill=red)
)
ws.conditional_formatting.add(
    v_range, FormulaRule(formula=["AND(ISNUMBER(V2),V2<-18)"], fill=red)
)

# YELLOW: -18 <= value <= -15 (numbers only)
ws.conditional_formatting.add(
    u_range, FormulaRule(formula=["AND(ISNUMBER(U2),U2>=-18,U2<=-15)"], fill=yellow)
)
ws.conditional_formatting.add(
    v_range, FormulaRule(formula=["AND(ISNUMBER(V2),V2>=-18,V2<=-15)"], fill=yellow)
)

# --- Save ---
wb.save(reservation_file)
print("✅ Updated values and applied numeric conditional formatting.")  